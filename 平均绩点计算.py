from openpyxl import load_workbook
from openpyxl.cell import cell
from openpyxl.styles import Font

courseCredit = []  # 课程学分数
courseName = []  # 课程名称


# 读取xlsx的Sheet1工作表
workBook = load_workbook("grade.xlsx")
rawSheet = workBook.active
outSheet = workBook.create_sheet("中间数据")

maxCol = rawSheet.max_column  # 最大列数
maxRow = rawSheet.max_row  # 最大行数


# 对第一行进行遍历，获取科目名称、科目学分数
for i in range(2, maxCol+1):
    list = (rawSheet.cell(row=1, column=i).value).split("-")  # 这里可以修改分割标识符
    name = list[0]+"-"+list[1]
    credit = float(list[1])
    courseName.append(name)
    courseCredit.append(credit)
    outSheet.cell(row=1, column=i, value=name)


# 第二次对成绩进行遍历，转分数为绩点，并将姓名列进行提取
for i in range(2, maxRow+1):
    studentName = rawSheet.cell(row=i, column=1).value
    outSheet.cell(row=i, column=1, value=studentName)

    # begin for j
    for j in range(2, maxCol+1):
        grade = rawSheet.cell(row=i, column=j).value
        if (grade == "优"):
            outSheet.cell(row=i, column=j, value=4.5)
        elif (grade == "良"):
            outSheet.cell(row=i, column=j, value=3.5)
        elif (grade == "中"):
            outSheet.cell(row=i, column=j, value=2.5)
        elif (grade == "及格"):
            outSheet.cell(row=i, column=j, value=1.5)
        elif (grade == "不及格"):
            outSheet.cell(row=i, column=j, value=0)
        elif ((grade == "缓考") | (grade == "未选")):
            outSheet.cell(row=i, column=j, value=grade)
        else:
            if (int(grade) >= 60):
                outSheet.cell(row=i, column=j, value=float(grade) / 10 - 5)
            else:
                outSheet.cell(row=i, column=j, value=0)
    #end for j

    stuCredit=0
    stuAveGradePoint=0
    for j in range(2,maxCol+1):
        stuGradePoint=outSheet.cell(row=i,column=j).value
        if((stuGradePoint!="缓考")&(stuGradePoint!="未选")):
            stuCredit=stuCredit+courseCredit[j-2]
            outSheet.cell(row=i,column=maxCol+1,value=stuCredit)
            if(float(stuGradePoint)>0):
                stuAveGradePoint=stuAveGradePoint+courseCredit[j-2]*float(stuGradePoint)
                outSheet.cell(row=i,column=maxCol+2,value=stuAveGradePoint/stuCredit)
    rawSheet.cell(row=i,column=maxCol+1,value=stuAveGradePoint/stuCredit)

#添加表头，保存文件   
rawSheet.cell(row=1,column=maxCol+1,value="平均绩点")  
outSheet.cell(row = 1, column = 1, value = "姓名")
outSheet.cell(row=1,column=maxCol+1,value="总学分")
outSheet.cell(row=1,column=maxCol+2,value="平均绩点")
workBook.save("平均绩点.xlsx")
